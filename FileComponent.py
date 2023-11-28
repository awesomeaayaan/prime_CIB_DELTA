from qrlib.QRComponent import QRComponent
from robot.libraries.BuiltIn import BuiltIn
from Errors import DataNotFoundError
from datetime import datetime
from EmailComponent import EmailComponent
from DBComponent import DBComponent
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
import ast
import os
import re
import time
import Utils
import sqlite3
import Constants
import csv
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime

display = BuiltIn().log_to_console

def cal_weights(row, *args):
    return float(row) * float(args)
 
class FileComponent(QRComponent):
    def __init__(self):
        super().__init__()
        self.dataframe = pd.DataFrame()
        self.db = DBComponent()
        self.email = EmailComponent()

    def set_dataframe(self, df: pd.DataFrame = pd.DataFrame()):
        self.dataframe = df
        # display(f'DATAFRAME FROM FILE COMPONENT{df.columns}')
    def extract_today_date():
        current_datetime = datetime.now()
        formatted_date = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
        return formatted_date

    def collect_user_reqiured_data(self, user_data: dict) -> dict:#handle cib black list incremental data
        # columns = user_data.keys()
       
        def filter_name(name: str) -> str:
            # * Regex to find correct name using regex pattern(Only for name)
            res = re.match("[a-zA-Z .()_\\\/&'+-]*", name)
            if res and (name.find('None') < 0):
                name = res.group()
            else:
                name = ''
            return name.strip()

        logger = self.run_item.logger
        logger.info('Combining personal information name')
        columns = user_data.keys()
        search_list = []

        account_nature = [i for i in user_data.keys() if i.find('Account_Nature') >= 0]
        account_nature = str(user_data[account_nature[0]]).lower().strip()
        logger.info(f'Nature of account is {account_nature}')
        if account_nature == 'individual':
            name = user_data.get('Name')
            search_list.append('name') if name and (user_data['Account_Nature'] == 'Individual') else None
            # display(f'this is search list{search_list}')

            father_name = user_data.get('FatherName')
            search_list.append('fathername') if father_name and (user_data['Account_Nature'] == 'Individual') else None

            # dob = [str(user_data[i]) for i in columns if i.find('DOB')>=0]
            date_of_birth = user_data.get('DOB')
            parsed_date = datetime.strptime(date_of_birth,"%Y-%m-%d")
            dob = parsed_date.strftime("%d/%m/%Y")
            search_list.append('dob') if dob and (user_data['Account_Nature'] == 'Individual') else None


            citizenship = [str(user_data[i]) for i in columns if i.find('CitizenshipDetails')>=0]
            # citizenship_no = []
            # display(f'citizenship===>{citizenship}')
            for item in citizenship:
                citizen_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    citizen_list = ast.literal_eval(item)
                    if isinstance(citizen_list, list):
                        for data in citizen_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                citizen_number = data.split('|')[0]
                                display(f'CITIZEN_NUMBER FROM FILE COMPONENT==>{citizen_number}')
                                # citizenship_no.append(citizen_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                                search_list.append('citizenship_no') if citizen_number and (user_data['Account_Nature'] == 'Individual') else None

                                return {
                                        'account_nature':account_nature,
                                        'name':str(name).strip(),
                                        'fathername':str(father_name).strip(),
                                        'dob':str(dob).strip(),
                                        'citizenship_no':citizen_number,
                                        # 'gfather_name':x_grandfather_name.strip(),
                                        # 'citizenship_no':str(citizenship).strip() if citizenship else '',
                                        # 'unique_id':str(user_data['unique_id']).strip(),
                                        # 'action':str(user_data['action']).strip(),
                                        # 'pan_no':pan_no if pan_no else '',
                                        # 'registration_no':str(registartion_no).strip() if registartion_no else '',
                                        # 'reg_no': reg_no if reg_no else '',
                                        # 'account_no': str(account_no).strip() if account_no else '',
                                        # 'main_code':user_data['main_code'],
                                        # 'client_code':user_data['client_code'],
                                        'search_list':search_list,
                                    }
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        citizen_number = item.split('|')[0]
                        # citizenship_no.append(citizen_number)
                        # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                        search_list.append('citizenship_no') if citizen_number and (user_data['Account_Nature'] == 'Individual') else None
                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'citizenship_no':citizen_number,
                            # 'gfather_name':x_grandfather_name.strip(),
                            # 'citizenship_no':str(citizenship).strip() if citizenship else '',
                            # 'unique_id':str(user_data['unique_id']).strip(),
                            # 'action':str(user_data['action']).strip(),
                            # 'pan_no':pan_no if pan_no else '',
                            # 'registration_no':str(registartion_no).strip() if registartion_no else '',
                            # 'reg_no': reg_no if reg_no else '',
                            # 'account_no': str(account_no).strip() if account_no else '',
                            # 'main_code':user_data['main_code'],
                            # 'client_code':user_data['client_code'],
                            'search_list':search_list,
                        }
                    else:
                        print(f"Invalid format in item: {item}")
            pan = [str(user_data[i]) for i in columns if i.find('PANDetails')>=0]
            for item in pan:
                pan_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    pan_list = ast.literal_eval(item)
                    if isinstance(pan_list, list):
                        for data in pan_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                pan_number = data.split('|')[0]
                                display(f'pan_NUMBER FROM FILE COMPONENT==>{pan_number}')
                                # citizenship_no.append(pan_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = pan_number
                                search_list.append('pan_no') if pan_number and (user_data['Account_Nature'] == 'Individual') else None

                                if 'citizenship_no' in search_list:
                                    search_list.remove('citizenship_no')

                                return {
                                        'account_nature':account_nature,
                                        'name':str(name).strip(),
                                        'fathername':str(father_name).strip(),
                                        'dob':str(dob).strip(),
                                        'pan_no':pan_number if pan_number else '',
                                        'search_list':search_list,
                                    }
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        pan_number = item.split('|')[0]
                        display(f'PANNUMBER===================================>{pan_number}')
                        # incremental_df.at[index, 'CitizenshipNumber'] = pan_number
                        search_list.append('pan_no') if pan_number and (user_data['Account_Nature'] == 'Individual') else None
                        if 'citizenship_no' in search_list:
                            search_list.remove('citizenship_no')

                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'pan_no':pan_number if pan_number else '',
                            'search_list':search_list,
                        }
                    
            passport = [str(user_data[i]) for i in columns if i.find('PassportDetails')>=0]
            for item in passport:
                passport_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    passport_list = ast.literal_eval(item)
                    if isinstance(passport_list, list):
                        for data in passport_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                passport_number = data.split('|')[0]
                                display(f'passport_NUMBER FROM FILE COMPONENT==>{passport_number}')
                                # citizenship_no.append(passport_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = passport_number
                                search_list.append('passport_no') if passport_number and (user_data['Account_Nature'] == 'Individual') else None

                                if 'citizenship_no' in search_list:
                                    search_list.remove('citizenship_no')

                                return {
                                        'account_nature':account_nature,
                                        'name':str(name).strip(),
                                        'fathername':str(father_name).strip(),
                                        'dob':str(dob).strip(),
                                        'passport_no':passport_number if passport_number else '',
                                        'search_list':search_list,
                                    }
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        passport_number = item.split('|')[0]
                        display(f'passportNUMBER===================================>{passport_number}')
                        # incremental_df.at[index, 'CitizenshipNumber'] = passport_number
                        search_list.append('passport_no') if passport_number and (user_data['Account_Nature'] == 'Individual') else None
                        if 'citizenship_no' in search_list:
                            search_list.remove('citizenship_no')

                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'passport_no':passport_number if passport_number else '',
                            'search_list':search_list,
                        }
                    
            drivinglisence = [str(user_data[i]) for i in columns if i.find('DrivingLicenseDetails')>=0]
            for item in drivinglisence:
                drivinglisence_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    drivinglisence_list = ast.literal_eval(item)
                    if isinstance(drivinglisence_list, list):
                        for data in drivinglisence_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                drivinglisence_number = data.split('|')[0]
                                display(f'drivinglisence_NUMBER FROM FILE COMPONENT==>{drivinglisence_number}')
                                # citizenship_no.append(drivinglisence_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = drivinglisence_number
                                search_list.append('drivinglisence_no') if drivinglisence_number and (user_data['Account_Nature'] == 'Individual') else None

                                if 'citizenship_no' in search_list:
                                    search_list.remove('citizenship_no')

                                return {
                                        'account_nature':account_nature,
                                        'name':str(name).strip(),
                                        'fathername':str(father_name).strip(),
                                        'dob':str(dob).strip(),
                                        'drivinglisence_no':drivinglisence_number if drivinglisence_number else '',
                                        'search_list':search_list,
                                    }
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        drivinglisence_number = item.split('|')[0]
                        display(f'drivinglisenceNUMBER===================================>{drivinglisence_number}')
                        # incremental_df.at[index, 'CitizenshipNumber'] = drivinglisence_number
                        search_list.append('drivinglisence_no') if drivinglisence_number and (user_data['Account_Nature'] == 'Individual') else None
                        if 'citizenship_no' in search_list:
                            search_list.remove('citizenship_no')

                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'drivinglisence_no':drivinglisence_number if drivinglisence_number else '',
                            'search_list':search_list,
                        }
            voterid = [str(user_data[i]) for i in columns if i.find('VoterIDDetails')>=0]
            for item in voterid:
                voterid_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    voterid_list = ast.literal_eval(item)
                    if isinstance(voterid_list, list):
                        for data in voterid_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                voterid_number = data.split('|')[0]
                                display(f'voterid_NUMBER FROM FILE COMPONENT==>{voterid_number}')
                                # citizenship_no.append(voterid_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = voterid_number
                                search_list.append('voterid_no') if voterid_number and (user_data['Account_Nature'] == 'Individual') else None

                                if 'citizenship_no' in search_list:
                                    search_list.remove('citizenship_no')

                                return {
                                        'account_nature':account_nature,
                                        'name':str(name).strip(),
                                        'fathername':str(father_name).strip(),
                                        'dob':str(dob).strip(),
                                        'voterid_no':voterid_number if voterid_number else '',
                                        'search_list':search_list,
                                    }
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        voterid_number = item.split('|')[0]
                        display(f'voteridNUMBER===================================>{voterid_number}')
                        # incremental_df.at[index, 'CitizenshipNumber'] = voterid_number
                        search_list.append('voterid_no') if voterid_number and (user_data['Account_Nature'] == 'Individual') else None
                        if 'citizenship_no' in search_list:
                            search_list.remove('citizenship_no')

                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'voterid_no':voterid_number if voterid_number else '',
                            'search_list':search_list,
                        } 
            indianembassy = [str(user_data[i]) for i in columns if i.find('IndianEmbassyDetails')>=0]
            for item in indianembassy:
                indianembassy_number = ""
                item = str(item)
                if '|' in item:
                        indianembassy_number = item.split('|')[0]
                        display(f'indianembassyNUMBER===================================>{indianembassy_number}')
                        # incremental_df.at[index, 'CitizenshipNumber'] = indianembassy_number
                        search_list.append('indianembassy_no') if indianembassy_number and (user_data['Account_Nature'] == 'Individual') else None
                        if 'citizenship_no' in search_list:
                            search_list.remove('citizenship_no')

                        return {
                            'account_nature':account_nature,
                            'name':str(name).strip(),
                            'fathername':str(father_name).strip(),
                            'dob':str(dob).strip(),
                            'indianembassy_no':indianembassy_number if indianembassy_number else '',
                            'search_list':search_list,
                        }    
                     
        if account_nature == 'institutions':
            name = user_data.get('Name')
            search_list.append('Name') if name else ''
            pan = [str(user_data[i]) for i in columns if i.find('PANDetails')>=0]
            pan_no = []
            # display(f'pan===>{pan}')
            for item in pan:
                pan_number = ""
                item = str(item)
                # Check if the item is a string or a list
                if ',' in item:
                    pan_list = ast.literal_eval(item)
                    if isinstance(pan_list, list):
                        for data in pan_list:
                            # Check if '|' exists before splitting
                            if '|' in data:
                                pan_number = data.split('|')[0]
                                pan_no.append(pan_number)
                                # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                                # search_list.append('pan_no') if pan_number and (user_data['Account_Nature'] == 'institutions') else None
                            else:
                                print(f"Invalid format in item: {data}")
                else:
                    # Check if '|' exists before splitting
                    if '|' in item:
                        pan_number = item.split('|')[0]
                        pan_no.append(pan_number)
                        # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                        # search_list.append('pan_no') if pan_number and (user_data['Account_Nature'] == 'Individual') else None
                    else:
                        print(f"Invalid format in item: {item}")


            search_list.append('pan_no') if pan_number and (user_data['Account_Nature'] == 'Institutions') else None

            # registartion_no = [str(user_data[i]) for i in columns if i.find('CompanyDetails')>=0]
            registartion_no = [str(user_data[i]) for i in columns if i.find('CompanyDetails')>=0]
            display(f'COMPANY DETAIL============>{registartion_no}')
            for item in registartion_no:
                reg_number = ""
                item = str(item)
                display(f'THIS IS COMPANY DETAILS{item}')
                reg_no = []
                if '|' in item:
                        reg_number = item.split('|')[0]
                        display(f'THIS IS REGISTRATION NUMBER ELSE{reg_number}')
                        # print(reg_number)
                        reg_no.append(reg_number)
                        # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                else:
                    print(f"Invalid format in item: {item}")
                # if ',' in item:
                #     reg_list = ast.literal_eval(item)
                #     if isinstance(reg_list, list):
                #         for data in reg_list:
                #             # Check if '|' exists before splitting
                #             if '|' in data:
                #                 reg_number = data.split('|')[0]
                #                 display(f'THIS IS REGISTRATION NUMBER IF {reg_number}')
                #                 reg_no.append(reg_number)
                #                 # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                #             else:
                #                 print(f"Invalid format in item: {data}")
                # else:
                #     # Check if '|' exists before splitting
                #     if '|' in item:
                #         reg_number = item.split('|')[0]
                #         display(f'THIS IS REGISTRATION NUMBER ELSE{reg_number}')
                #         # print(reg_number)
                #         reg_no.append(reg_number)
                #         # incremental_df.at[index, 'CitizenshipNumber'] = citizen_number
                #     else:
                #         print(f"Invalid format in item: {item}")
            search_list.append('registration_no') if reg_number  else ''
                        
            return {
                    'account_nature':account_nature,
                    'name':str(name).strip(),
                    'fathername':'',
                    'dob':'',
                    'citizenship_no': '',
                    # 'gfather_name':x_grandfather_name.strip(),
                    # 'citizenship_no':str(citizenship).strip() if citizenship else '',
                    # 'unique_id':str(user_data['unique_id']).strip(),
                    # 'action':str(user_data['action']).strip(),
                    'pan_no':pan_number if pan_number else '',
                    # 'registration_no':str(registartion_no).strip() if registartion_no else '',
                    'registration_no': reg_number if reg_number else '',
                    # 'account_no': str(account_no).strip() if account_no else '',
                    # 'main_code':user_data['main_code'],
                    # 'client_code':user_data['client_code'],
                    'search_list':search_list,
                                }         

        
    def merge_columns(self, columns: list[str], new_column: str):
        logger = self.run_item.logger
        logger.info('joining multiple columns')
        self.dataframe = self.dataframe.fillna('')
        self.dataframe[new_column] = self.dataframe[columns].apply(" ".join, axis=1)
        logger.info('joining columns')   

    def calculate_weightage(self, column: str, value: int, by_match: bool=True,  df: pd.DataFrame = pd.DataFrame(), **kwargs):
        '''
            Required kwargs value:
                new_column: string
                match_item: string
                soundex: boolean
                weight_col: string
        '''
        start_time = time.time()
        logger = self.run_item.logger
        weight_col = kwargs['weight_col']
        float_weight = float(value/100)
        logger.info(f'Columns : {self.dataframe.columns.to_list()}')

        required_data = ['new_column', 'match_item', 'soundex']
        if not all(x in kwargs.keys() for x in required_data):
            raise Exception(f'{required_data} datas are required')
        
        new_column = str(kwargs['new_column'])

        if df.empty:
           was_empty = True
           df = self.dataframe
        else:
            was_empty = False

        if by_match:
            soundex: bool = kwargs['soundex']
            df[weight_col] = df[column].apply(
                Utils.compare_two_stirngs_in_df,
                string2 = str(kwargs['match_item']),
                soundex = soundex
            )
            column = weight_col
    
        logger.info(f'Making sure {column} column is int or float data type')
        df[column] = pd.to_numeric(df[column], errors='coerce').fillna(0)
        df[new_column] = df[column] * float_weight
        
        logger.info(f'Dataframe shapt = {df.shape}')
        if was_empty:
            self.dataframe = df
        finish_time = time.time()
        logger.info(f'Time consumed for weightage calculation of {column} is {int(finish_time - start_time)}.')
        logger.info(f"Weightage dataframe columns: {df.columns}")
        return df

    def combine_columns_by_addition(self,columns: list[str], new_col_name: str='similarity'):#hereichange
        logger = self.run_item.logger
        logger.info('Add up the values of each column')
        logger.info(self.dataframe)
        column_sum = self.dataframe[columns].sum(axis=1)
        logger.info('Create a new column with the sum values')
        self.dataframe[new_col_name] = column_sum
        result_cib = self.dataframe
        display(f'This is the column ==> {result_cib.columns}')
        # if result_cib['similarity'] <= 90.00:
        # result = result_cib[result_cib['similarity'] <= 90.00]
        # result.drop(columns=[],inplace=True)
        # self.db.insert_report_not_matched_with_cib_data(result)
        # res = result_cib[result_cib['total'] > 90.000]
        # res.to_excel('cib_result.xlsx')
        # print(self.dataframe)
        # display(F'FILE COMPONENT DATAFRAME FROM COMBINE COLUMN BY ADDITION{self.dataframe.columns}')

    def get_data_by_value_limit(self, column: str,
                                value: float or int,
                                required_columns: list[str],
                                add_colum: bool = False,
                                column_data: dict = {}) -> list[dict]:
        '''
        key value pair in column_data parameter:
            name: str, name of column going to be added
            values: Any, value of data to be inserted
        '''
        logger = self.run_item.logger
        display(f'required columns = {required_columns}')
        query = f'{column} > {value}'
        logger.info('==========Get data by value limi is called==========')
        required_columns.append(column)
        # try:
        #     res = self.dataframe[self.dataframe[column] == 100.0]
        #     if res.shape[0] == 1:
        #         logger.info('Single data found from data frame.')
        #         res = res[required_columns]
        #         if add_colum and column_data:
        #             column_name = column_data['name']
        #             column_values = column_data['values']
        #             res.insert(0, column_name, column_values)
        #         res = res.to_dict(orient='records')
        #         return res
        #     del res
        #     logger.info('Data with 100 percent match did not found.')
        # except Exception as e:
        #     logger.info(e)
        # if self.dataframe[self.dataframe[column] > float(value)]:
        result = self.dataframe[self.dataframe[column] > float(value)]
        # if result is not None:
        # display(f'RESULT COLUMN ARE == > {result.columns}')
        logger.info('Dataframe query successfull.')
        display(f'Dataframe columns = {self.dataframe.columns.to_list()}')
        result = result[required_columns]
        # display(f'RESULT COLUMN ARE == > {result.columns}')
        logger.info(f'add_column = {add_colum}, column_data = {column_data}')
        if add_colum and column_data:
            column_name = column_data['name']
            column_values = column_data['values']
            result.insert(0, column_name, column_values)
        display(f'result columns1 = {result.columns.to_list()}')
        if 'maincode' in result.columns.to_list():
            result.rename(columns={'maincode':'account_no'}, inplace=True)
        display(f'result columns2 = {result.columns.to_list()}')
        result  = result.to_dict(orient='records')
        logger.info('List of dictionary ready from dataframe.')
        return result
        # elif result is None:
        #     result_df = self.dataframe[self.dataframe[column] <= float(value)]
        #     self.db.insert_report_not_matched_with_cib_data(result_df)
    
    def insert_column_at_index(self, index: int, column: str, value: str):
        logger = self.run_item.logger
        column = column.title()
        # display(f'THIS IS FROM INSERT COLUMN AT INDEX:  {column}====>{value}')
        self.dataframe.insert(index, column, value)
        logger.info('Column insertion successful')

    def get_number_of_row_data(self, df=pd.DataFrame()) -> int:
        return self.dataframe.shape[0]
    
    def get_columns_name_of_sheet(self) -> list:
        return self.dataframe.columns.to_list()
    
    def add_new_columns_to_datasheet(self, col_name: str, values, length: int = 0):
        '''
            Must contains total number of values
        '''
        logger = self.run_item.logger
        logger.info(f'Given input length is {length}')
        display(F'given input length is {length}')
        if not isinstance(values, list):
            values = [str(values) for _ in range(length)]
        display(self.dataframe.shape[0])
        display(f'Value is {values}')
        display(f'Value length is {len(values)}')
        # if self.dataframe.shape[0] != len(values):
        #     logger.error(f'Values length is not equals to {self.dataframe.shape[0]}')
        #     return False
        logger.info('Adding new columns with values')
        if len(values) >1:
            # values = [item + '&' for item in values]
            values = [' '.join(values)]
            display(f'List value is ==> {values}')
            self.dataframe[col_name] = values
        else:
            values = values
            self.dataframe[col_name] = values
        return True
    def collect_maincode_data(self, other_df: pd.DataFrame):
        logger = self.run_item.logger
        # display(f'COLUMS_OF_OTHER_DF===> {other_df.columns}')
        # display(f'DATAFRAME_COLUMN==>{self.dataframe.columns}')
        # display(f'DATAFRAME COLUMNS ==> {self.dataframe.columns}')
        # main_code = self.dataframe['maincode'].to_list()
        # display(f'MAINCODE==>{main_code}')
        # main_code = other_df['maincode'].to_list()
        # main_code =  [str(x).zfill(Constants.MAINCODE_LENGTH) for x in main_code]
        # display(f'MAIN CODE IS {main_code}')
        # main_code =  [str(x).zfill(Constants.MAINCODE_LENGTH) for x in main_code]
        # result = other_df[other_df['main_code'].isin(main_code)][['clientcode', 'maincode', 'branchcode', 'isblocked', 'name']]
        # self.dataframe = pd.merge(self.dataframe,result,left_on='maincode', right_on='maincode',how='inner')
        client_code = self.dataframe['clientcode'].to_list()
        # display(f'CLIENT_CODE==>{client_code}')
        client_code = [str(x).zfill(Constants.CLIENTCODE_LENGTH) for x in client_code]
        result = other_df[other_df['clientcode'].isin(client_code)][['clientcode', 'maincode', 'branchcode', 'isblocked', 'name','actypedesc']]
        # display(f'RESULT_COLUMN ARE ==> {result.columns}')
        # result = other_df[other_df['maincode'].isin(main_code)][['clientcode', 'maincode', 'branchcode', 'isblocked', 'name']]
        result.rename(columns={'name':'Account Holder'}, inplace=True)
        logger.info(f'result column {result.columns.to_list()}')
        logger.info('clientcode and maincode from client master retirved.')
        # self.dataframe.merge(result, on='clientcode', how='inner')
        # self.dataframe = pd.merge(self.dataframe,result,left_on='maincode', right_on='maincode',how='inner')
        self.dataframe = pd.merge(self.dataframe, result, left_on='clientcode', right_on='clientcode', how='inner')
        # display(f'DATAFRAME COLUMNS ==> {self.dataframe.columns}')
        # self.dataframe.drop(columns=['name_weight','fathersname_weight','nepdate_weight','citizenshipno_weight','search_type','Account Holder'],axis=1,inplace=True)
        # self.dataframe.drop(columns=['search_type','Account Holder'],axis=1,inplace=True)
        if 'fathersname_weight' in self.dataframe.columns:
            if 'citizenshipno_weight' in self.dataframe.columns:
                self.dataframe.drop(columns=['name_weight','fathersname_weight','nepdate_weight','citizenshipno_weight','search_type','Account Holder'],axis=1,inplace=True)
                self.remove_columns(self.dataframe.columns)
                column_rename = {
                    'name_percent':'Name Match %',
                    'nepdate_percent':'DOB Match %',
                    'fathersname_percent': 'Father Name %',
                    'citizenshipno_percent':'Citizenship Match %',
                    'clientcode':'CBS Client Code',
                    'name':'CBS Account Name',
                    'fathersname':'CBS Father Name',
                    'citizenshipno' : 'CBS Citizenship No',
                    'nepdate' : 'CBS Date of birth(BS)',
                    'maincode' : 'CBS Account Number',
                    'branchcode': 'Branch Code',
                    'actypedesc': 'Account Description',
                    'similarity' : 'Total Similarity %',
                    'dlicenceidno' : 'Driving Licence No',
                    'indianembassyregno': 'Indian Embassy Reg No',
                    'pannumber': 'PAN No',
                    'passportno' : 'Passport No'
                }
                self.dataframe.rename(columns=column_rename,inplace=True)
                self.change_status_column_value()
                display(self.dataframe.head())
                display(self.dataframe.columns)
                
                # self.dataframe.to_excel(Constants.FINAL_REPORT,index=False)
                # self.dataframe = self.dataframe.reindex(sorted(self.dataframe.columns),axis=1)#'CBS Account Status'
                desired_column = ['Branch Code','CBS Client Code','CBS Account Number','Account Description','CBS Account Status','CBS Account Name','Name Match %','CBS Father Name','Father Name %','CBS Date of birth(BS)','DOB Match %','CBS Citizenship No','Citizenship Match %','Total Similarity %','Result of CIB Screening','CIB Name','CIB Father Name','CIB DOB','CIB Citizenship No','CIB Gender','CIB Black Listed No','CIB Black Listed Date','Driving Licence No','Indian Embassy Reg No','PAN No','Passport No']
                self.dataframe = self.dataframe[desired_column]
            elif  'pannumber_weight' in self.dataframe.columns:
                self.dataframe.drop(columns=['name_weight','fathersname_weight','nepdate_weight','pannumber_weight','search_type','Account Holder'],axis=1,inplace=True)
                self.remove_columns(self.dataframe.columns)
                column_rename = {
                    'name_percent':'Name Match %',
                    'nepdate_percent':'DOB Match %',
                    'fathersname_percent': 'Father Name %',
                    # 'citizenshipno_percent':'Citizenship Match %',
                    'pannumber_percent': 'PAN NUMBER %',
                    'clientcode':'CBS Client Code',
                    'name':'CBS Account Name',
                    'fathersname':'CBS Father Name',
                    # 'pannumber': 'PAN NUMBER',
                    # 'citizenshipno' : 'CBS Citizenship No',
                    'nepdate' : 'CBS Date of birth(BS)',
                    'maincode' : 'CBS Account Number',
                    'branchcode': 'Branch Code',
                    'actypedesc': 'Account Description',
                    'similarity' : 'Total Similarity %',
                    'dlicenceidno' : 'Driving Licence No',
                    'indianembassyregno': 'Indian Embassy Reg No',
                    'pannumber': 'PAN No',
                    'passportno' : 'Passport No'
                }
                self.dataframe.rename(columns=column_rename,inplace=True)
                self.change_status_column_value()
                display(self.dataframe.head())
                display(self.dataframe.columns)
                
                # self.dataframe.to_excel(Constants.FINAL_REPORT,index=False)
                # self.dataframe = self.dataframe.reindex(sorted(self.dataframe.columns),axis=1)#'CBS Account Status'
                desired_column = ['Branch Code','CBS Client Code','CBS Account Number','Account Description','CBS Account Status','CBS Account Name','Name Match %','CBS Father Name','Father Name %','CBS Date of birth(BS)','DOB Match %','PAN No','PAN NUMBER %','Total Similarity %','Result of CIB Screening','CIB Name','CIB Father Name','CIB DOB','CIB PAN NO','CIB Gender','CIB Black Listed No','CIB Black Listed Date','Driving Licence No','Indian Embassy Reg No','PAN No','Passport No']
                self.dataframe = self.dataframe[desired_column]
        if 'pannumber_weight' and 'comregnum_weight' in self.dataframe.columns:
              display(f'Column of company details are: {self.dataframe.columns}')
              self.dataframe.drop(columns=['name_weight','pannumber_weight','comregnum_weight','search_type','Account Holder'],axis=1,inplace=True)
              self.remove_columns(self.dataframe.columns)
            #   self.change_status_column_value()
              column_rename = {
                    'clientcode':'CBS Client Code',
                    'name':'CBS Account Name',
                    'name_percent':'Name Match %',
                    # 'clientstatus' : 'CBS Account Status',
                    # 'fathersname':'CBS Father Name',
                    # 'citizenshipno' : 'CBS Citizenship No',
                    # 'nepdate' : 'CBS Date of birth(BS)',
                    'maincode' : 'CBS Account Number',
                    'branchcode': 'Branch Code',
                    'actypedesc': 'Account Description',
                    'similarity' : 'Total Similarity %',
                    # 'dlicenceidno' : 'Driving Licence No',
                    # 'indianembassyregno': 'Indian Embassy Reg No',
                    'pannumber': 'PAN No',
                    'comregnum' : 'ComRegister No',
                    'comregnum_percent': 'Company RegMatch %',
                    'pannumber_percent':'PAN Match %'

                }
              self.dataframe.rename(columns=column_rename,inplace=True)
              self.change_status_column_value()
              display(f'{self.dataframe.head()}')
              df = self.dataframe
              desired_column = ['Branch Code','CBS Client Code','CBS Account Number','Account Description','CBS Account Status','CBS Account Name','Name Match %','ComRegister No','Company RegMatch %','PAN No','PAN Match %','Total Similarity %','Result of CIB Screening','CIB Name','CIB Com Registration No','CIB PAN NO','CIB Black Listed No','CIB Black Listed Date']
              self.dataframe = self.dataframe[desired_column]
                # self.db.insert_report(df)
                # display(self.dataframe.head())
                # self.change_status_column_value()
        
        # # self.dataframe.to_excel(Constants.FINAL_REPORT,index=False)
        # # self.dataframe = self.dataframe.reindex(sorted(self.dataframe.columns),axis=1)
        # desired_column = ['Branch Code','CBS Client Code','CBS Account Number','Account Description','CBS Account Status','CBS Account Name','Name Match %','CBS Father Name','Father Name %','DOB Match %','CBS Citizenship No','Citizenship Match %','Total Similarity %','Result of CIB Screening','CIB Name','CIB Father Name','CIB DOB','CIB Citizenship No','CIB Gender','CIB Blacklisted No','CIB Blacklisted Date','Driving Licence No','Indian Embassy Reg No','PAN No','Passport No']
        # self.dataframe = self.dataframe[desired_column]
        # display('Apple-------------------Apple')
        excel_writer = pd.ExcelWriter(Constants.FINAL_REPORT,engine='openpyxl')
        # excel_writer = pd.ExcelWriter(f'{Constants.REPORT_PATH}/',engine='openpyxl')
        self.dataframe.to_excel(excel_writer,index=False,sheet_name='Sheet1')

        workbook = excel_writer.book
        worksheet = excel_writer.sheets['Sheet1']
        result_of_CIB = ['Match with CIB Black List','No Match with CIB Black List']
        if 'Father Name %' in self.dataframe:
            df = self.dataframe
            dropdownrange = 'O2:O{}'.format(len(self.dataframe)+1)
            #define the column to be highlighted
            highlighted_column = 'O'
            data_validation = DataValidation(type="list",formula1='"{}"'.format(','.join(result_of_CIB)),allow_blank=True)
        
            worksheet.add_data_validation(data_validation)
            data_validation.add(dropdownrange)

            #define the range of the cell in thte column to be highlighted
            highlighted_range = '{}2:{}{}'.format(highlighted_column,highlighted_column,len(df)+1)

            #create a patternfill object to highlight tje cells (eg.yellow fill)
            highlight_fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

            #apply the highlight style to the specified column
            for row in worksheet[highlighted_range]:
                for cell in row:
                    cell.fill = highlight_fill
            
            #Save the Excel file 
            excel_writer.close()

             # self.dataframe.to_excel('file.xlsx',index=False)
            display("Dropdown list added to the cib result column in the excel file")

            display('FILE DOWNLOADED')
            self.email._authmail_and_send()
            
        if 'Company RegMatch %' in self.dataframe.columns:
            df = self.dataframe
            dropdownrange = 'M2:M{}'.format(len(self.dataframe)+1)
            highlighted_column = 'M'
            data_validation = DataValidation(type="list",formula1='"{}"'.format(','.join(result_of_CIB)),allow_blank=True)
            
            worksheet.add_data_validation(data_validation)
            data_validation.add(dropdownrange)

            #define the range of the cell in thte column to be highlighted
            highlighted_range = '{}2:{}{}'.format(highlighted_column,highlighted_column,len(df)+1)

            #create a patternfill object to highlight tje cells (eg.yellow fill)
            highlight_fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

            #apply the highlight style to the specified column
            for row in worksheet[highlighted_range]:
                for cell in row:
                    cell.fill = highlight_fill
            
            #Save the Excel file 
            excel_writer.close()
            

            # self.dataframe.to_excel('file.xlsx',index=False)
            display("Dropdown list added to the cib result column in the excel file")

            display('FILE DOWNLOADED')
            self.email._authmail_and_send()
            

    def read_file_and_send_email(self):
        # directory = r"C:\Users\RPA\Documents\RPA\CIB_BOT_Delta-master\Documents\test_email\attachments"
        directory = os.path.join(Constants.COMMON_PATH,'test_email','attachments')
        # current_path = os.getcwd()
        # directory = os.path.abspath(os.path.join(current_path,'..','..','..','Documents','test_email','attachments'))
        display(os.listdir(directory))
        for file in os.listdir(directory):
            if file.endswith('.xlsx'):
                display(f'File is ==> {file}')
                df = pd.read_excel(os.path.join(directory,file))
                if 'CBS Father Name' in df:
                    if df['Result of CIB Screening'].iloc[0] == 'Match with CIB Black List' or df['Result of CIB Screening'].iloc[0] == 'No Match with CIB Black List' and 'CBS Father Name' in df:
                        df['CBS Account Number'] = df['CBS Account Number'].astype(str)
                        display(df['Result of CIB Screening'].iloc[0])
                        self.db.insert_report(df)
                        file_path = os.path.join(directory,file)
                        os.remove(file_path)
                elif 'CIB Com Registration No' in df:
                    if df['Result of CIB Screening'].iloc[0] == 'Match with CIB Black List' or df['Result of CIB Screening'].iloc[0] == 'No Match with CIB Black List' and 'CIB Com Registration No' in df:
                        display(df['Result of CIB Screening'].iloc[0])
                        df['CBS Account Number'] = df['CBS Account Number'].astype(str)
                        self.db.insert_institution_report(df) 
                        file_path = os.path.join(directory,file)
                        os.remove(file_path)
                elif df['Result of CIB Screening'].iloc[0] == 'none':
                    df = pd.read_excel(os.path.join(directory,file))
                    branch_value = str(df['Branch Code'].iloc[0])
                    file_path = os.path.join(directory,file)
                    display(file_path)
                    display(branch_value)
                    display(df['Result of CIB Screening'].iloc[0])
                    self.email.follow_up_mail(branch_value,file_path)
                    os.remove(file_path)
    
    def clear_directory(self):
        pass
    def change_status_column_value(self, col_name: str = 'isblocked', length: int = 0):
        logger = self.run_item.logger
        logger.info(f'Given input length is {length}')
        columns = self.dataframe.columns.to_list()
        if col_name in columns:
            self.dataframe.rename(columns={col_name:'CBS Account Status'}, inplace=True)
            self.dataframe['CBS Account Status'] = self.dataframe['CBS Account Status'].apply(lambda x : Constants.REVRSE_RESTRICTION[str(x).strip().lower()]
                if str(x).strip().lower() in Constants.REVRSE_RESTRICTION.keys() else 'null')
            

    def remove_columns(self, columns: list, df = pd.DataFrame()) -> pd.DataFrame:
        logger = self.run_item.logger
        was_empty = False
        logger.info(f'Columns are {columns}')
        display(f'COLUMNS OF THE {columns}')
        if df.empty:
            df = self.dataframe
            was_empty = True
        logger.info('Removing columns from dataframe.')
        # self.dataframe.drop(axis=1, columns=columns,inplace=True)
        display(f'Succesfully remove the columns')
        df.drop(axis=1, columns=columns)
        logger.info('Removing data from dataframe sucessfull.')
        
        if was_empty:
            self.dataframe = df
        
        return df
    
    def convert_list_into_multiple_row(self,df):
        new_columns = []
        for i, data in df.iterrows():
            if data['Account_Nature'] == 'Institutions':
                # print(type(data))
                if 'PANDetails' in data:
                    if isinstance(data['PANDetails'], list):
                        for pan_element in data['PANDetails']:
                            if isinstance(data['CompanyDetails'],list):
                                for company_element in data['CompanyDetails']:
                                    new_data = {
                                        'Name': data['Name'],
                                        'DOB': data['DOB'],
                                        'Gender': data['Gender'],
                                        'FatherName': data['FatherName'],
                                        'CitizenshipDetails': data['CitizenshipDetails'],
                                        'PassportDetails': data['PassportDetails'],
                                        'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                        'VoterIDDetails': data['VoterIDDetails'],
                                        'PANDetails': pan_element,
                                        'CompanyDetails': company_element,
                                        'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                        'BlackLists': data['BlackLists'],
                                        'Account_Nature': data['Account_Nature'],
                                        'Status': data['Status']
                                    }
                                    new_columns.append(new_data)
                            elif isinstance(data['CompanyDetails'],str):
                                new_data = {
                                        'Name': data['Name'],
                                        'DOB': data['DOB'],
                                        'Gender': data['Gender'],
                                        'FatherName': data['FatherName'],
                                        'CitizenshipDetails': data['CitizenshipDetails'],
                                        'PassportDetails': data['PassportDetails'],
                                        'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                        'VoterIDDetails': data['VoterIDDetails'],
                                        'PANDetails': pan_element,
                                        'CompanyDetails': data['CompanyDetails'],
                                        'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                        'BlackLists': data['BlackLists'],
                                        'Account_Nature': data['Account_Nature'],
                                        'Status': data['Status']
                                    }
                                new_columns.append(new_data)
                    elif isinstance(data['PANDetails'], str):
                        if isinstance(data['CompanyDetails'],str):
                            new_data = {
                                            'Name': data['Name'],
                                            'DOB': data['DOB'],
                                            'Gender': data['Gender'],
                                            'FatherName': data['FatherName'],
                                            'CitizenshipDetails': data['CitizenshipDetails'],
                                            'PassportDetails': data['PassportDetails'],
                                            'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                            'VoterIDDetails': data['VoterIDDetails'],
                                            'PANDetails': data['PANDetails'],
                                            'CompanyDetails': data['CompanyDetails'],
                                            'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                            'BlackLists': data['BlackLists'],
                                            'Account_Nature': data['Account_Nature'],
                                            'Status': data['Status']
                                        }
                            new_columns.append(new_data)
                        elif isinstance(data['CompanyDetails'],list):
                            for company_element in data['CompanyDetails']:
                                    new_data = {
                                        'Name': data['Name'],
                                        'DOB': data['DOB'],
                                        'Gender': data['Gender'],
                                        'FatherName': data['FatherName'],
                                        'CitizenshipDetails': data['CitizenshipDetails'],
                                        'PassportDetails': data['PassportDetails'],
                                        'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                        'VoterIDDetails': data['VoterIDDetails'],
                                        'PANDetails': data['PANDetails'],
                                        'CompanyDetails': company_element,
                                        'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                        'BlackLists': data['BlackLists'],
                                        'Account_Nature': data['Account_Nature'],
                                        'Status': data['Status']
                                    }
                                    new_columns.append(new_data)

                else:
                    if isinstance(data['CompanyDetails'],str):
                            new_data = {
                                            'Name': data['Name'],
                                            'DOB': data['DOB'],
                                            'Gender': data['Gender'],
                                            'FatherName': data['FatherName'],
                                            'CitizenshipDetails': data['CitizenshipDetails'],
                                            'PassportDetails': data['PassportDetails'],
                                            'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                            'VoterIDDetails': data['VoterIDDetails'],
                                            'PANDetails': data['PANDetails'],
                                            'CompanyDetails': data['CompanyDetails'],
                                            'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                            'BlackLists': data['BlackLists'],
                                            'Account_Nature': data['Account_Nature'],
                                            'Status': data['Status']
                                        }
                            new_columns.append(new_data)
                    elif isinstance(data['CompanyDetails'],list):
                        for company_element in data['CompanyDetails']:
                                new_data = {
                                    'Name': data['Name'],
                                    'DOB': data['DOB'],
                                    'Gender': data['Gender'],
                                    'FatherName': data['FatherName'],
                                    'CitizenshipDetails': data['CitizenshipDetails'],
                                    'PassportDetails': data['PassportDetails'],
                                    'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                                    'VoterIDDetails': data['VoterIDDetails'],
                                    'PANDetails': data['PANDetails'],
                                    'CompanyDetails': company_element,
                                    'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                                    'BlackLists': data['BlackLists'],
                                    'Account_Nature': data['Account_Nature'],
                                    'Status': data['Status']
                                }
                                new_columns.append(new_data)
            elif data['Account_Nature'] == 'Individual':
                new_data = {
                    'Name': data['Name'],
                    'DOB': data['DOB'],
                    'Gender': data['Gender'],
                    'FatherName': data['FatherName'],
                    'CitizenshipDetails': data['CitizenshipDetails'],
                    'PassportDetails': data['PassportDetails'],
                    'DrivingLicenseDetails': data['DrivingLicenseDetails'],
                    'VoterIDDetails': data['VoterIDDetails'],
                    'PANDetails': data['PANDetails'],
                    'CompanyDetails': data['CompanyDetails'],
                    'IndianEmbassyDetails': data['IndianEmbassyDetails'],
                    'BlackLists': data['BlackLists'],
                    'Account_Nature': data['Account_Nature'],
                    'Status': data['Status']
                }
                new_columns.append(new_data)
                    # print(new_columns)
                # if isinstance(data[])

        new_df = pd.DataFrame(new_columns)
        no_duplicate_df = new_df[~new_df.duplicated(subset=new_df.columns.difference(['BlackLists']), keep='first')]
        print('Remove the duplicated data')
        return no_duplicate_df